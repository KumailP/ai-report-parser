import io
from app.logger import logger
from openpyxl import load_workbook
from fastapi import HTTPException

async def process_excel_file(report: io.BytesIO) -> dict:
    """
    Process an uploaded Excel file and extract data as rows.
    
    Args:
        report: The uploaded file object
        
    Returns:
        List of tuples containing cell values for each row
        
    Raises:
        HTTPException: For invalid file types or processing errors
    """
    try:
        logger.info(f"Starting Excel file processing: {report.filename}")
        
        # Validate file format
        if report.content_type not in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"]:
            logger.error(f"Invalid file type: {report.content_type}")
            raise HTTPException(status_code=400, detail="Invalid file type. Only .xlsx files are allowed.")
    
        # Read file into memory and load workbook
        xlsx = io.BytesIO(await report.read())
        wb = load_workbook(xlsx)

        if not wb.sheetnames:
            logger.error("Excel file has no sheets")
            raise HTTPException(status_code=400, detail="Excel file has no sheets")

        # Use the first sheet by default
        first_sheet = wb.sheetnames[0]
        logger.info(f"Processing sheet: {first_sheet}")
        ws = wb[first_sheet]

        # Extract all rows, replacing None with empty string
        rows = [
            tuple(cell.value if cell.value is not None else "" for cell in row) 
            for row in ws.iter_rows()
        ]

        
        logger.info(f"Successfully processed {len(rows)} rows from sheet '{first_sheet}'")
        
        return rows
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error processing Excel file {report.filename}: {str(e)}", exc_info=True)
        raise HTTPException(status_code=422, detail=f"Error processing Excel file: {str(e)}")